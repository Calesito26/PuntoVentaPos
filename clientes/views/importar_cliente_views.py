import openpyxl

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse

from clientes.models import Cliente


def importar_clientes(request):

    if request.method == 'POST':

        archivo = request.FILES.get('archivo')

        workbook = openpyxl.load_workbook(archivo)
        hoja = workbook.active

        for fila in hoja.iter_rows(min_row=2):

            nombre = fila[0].value
            tipo_documento = fila[1].value
            numero_documento = fila[2].value
            telefono = fila[3].value
            correo = fila[4].value
            direccion = fila[5].value

            Cliente.objects.create(
                nombre=nombre,
                tipo_documento=tipo_documento,
                numero_documento=numero_documento,
                telefono=telefono,
                correo=correo,
                direccion=direccion,
                activo=True
            )

        messages.success(
            request,
            'Clientes importados correctamente'
        )

        return redirect('clientes:cliente_list')

    return render(
        request,
        'clientes/importar_clientes.html'
    )


def descargar_ejemplo_clientes(request):

    workbook = openpyxl.Workbook()
    hoja = workbook.active

    hoja.append([
        'nombre',
        'tipo_documento',
        'numero_documento',
        'telefono',
        'correo',
        'direccion'
    ])

    hoja.append([
        'JUAN PEREZ',
        'DNI',
        '74413880',
        '999888777',
        'correo@gmail.com',
        'CHICLAYO'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=ejemplo_clientes.xlsx'
    )

    workbook.save(response)

    return response