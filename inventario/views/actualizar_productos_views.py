from decimal import Decimal

import openpyxl

from django.contrib import messages
from django.shortcuts import redirect
from django.shortcuts import render

from inventario.models import Sede
from inventario.models import StockBodega
from productos.models import Categoria
from productos.models import Producto
from usuarios.decorators import administrador_required


@administrador_required
def actualizar_productos_excel(request):

    if request.method == 'POST':

        archivo = request.FILES.get('archivo_excel')

        if not archivo:
            messages.error(
                request,
                'Debe seleccionar un archivo Excel.'
            )
            return redirect('inventario:actualizar_productos_excel')

        try:
            workbook = openpyxl.load_workbook(archivo)
            hoja = workbook.active

            actualizados = 0
            errores = []

            for fila in range(2, hoja.max_row + 1):

                codigo = hoja.cell(fila, 1).value
                nombre = hoja.cell(fila, 2).value
                descripcion = hoja.cell(fila, 3).value
                categoria_id = hoja.cell(fila, 4).value
                precio_compra = hoja.cell(fila, 5).value
                precio_venta = hoja.cell(fila, 6).value
                imagen_url = hoja.cell(fila, 7).value
                bodega_id = hoja.cell(fila, 8).value
                stock = hoja.cell(fila, 9).value
                stock_minimo = hoja.cell(fila, 10).value
                activo = hoja.cell(fila, 11).value
                es_insumo = hoja.cell(fila, 12).value
                receta_servicio_combo = hoja.cell(fila, 13).value

                if not codigo:
                    continue

                producto = Producto.objects.filter(
                    codigo=str(codigo).strip()
                ).first()

                if not producto:
                    errores.append(
                        f'Fila {fila}: producto con código {codigo} no existe.'
                    )
                    continue

                if nombre not in [None, '']:
                    producto.nombre = nombre

                if descripcion not in [None, '']:
                    producto.descripcion = descripcion

                if categoria_id not in [None, '']:
                    categoria = Categoria.objects.filter(
                        id=categoria_id
                    ).first()

                    if categoria:
                        producto.categoria = categoria
                    else:
                        errores.append(
                            f'Fila {fila}: categoría no existe.'
                        )

                if precio_compra not in [None, '']:
                    producto.precio_compra = Decimal(str(precio_compra))

                if precio_venta not in [None, '']:
                    producto.precio_venta = Decimal(str(precio_venta))

                if imagen_url not in [None, '']:
                    producto.imagen_url = imagen_url

                if activo not in [None, '']:
                    producto.activo = str(activo).strip().upper() not in [
                        'NO',
                        '0',
                        'FALSE',
                        'INACTIVO'
                    ]

                if es_insumo not in [None, '']:
                    producto.es_insumo = str(es_insumo).strip().upper() == 'SI'

                if receta_servicio_combo not in [None, '']:
                    receta_servicio_combo = str(
                        receta_servicio_combo
                    ).strip().upper()

                    if receta_servicio_combo not in [
                        'NINGUNO',
                        'RECETA',
                        'SERVICIO',
                        'COMBO'
                    ]:
                        receta_servicio_combo = 'NINGUNO'

                    producto.receta_servicio_combo = receta_servicio_combo

                producto.save()

                if bodega_id not in [None, '']:
                    sede = Sede.objects.filter(
                        id=bodega_id
                    ).first()

                    if sede:
                        stock_bodega, creado = StockBodega.objects.get_or_create(
                            sede=sede,
                            producto=producto,
                            defaults={
                                'stock': 0,
                                'stock_minimo': 5,
                                'activo': producto.activo
                            }
                        )

                        if stock not in [None, '']:
                            stock_bodega.stock = Decimal(str(stock))

                        if stock_minimo not in [None, '']:
                            stock_bodega.stock_minimo = Decimal(str(stock_minimo))

                        stock_bodega.activo = producto.activo
                        stock_bodega.save()
                    else:
                        errores.append(
                            f'Fila {fila}: bodega no existe.'
                        )

                actualizados += 1

            for error in errores:
                messages.warning(request, error)

            messages.success(
                request,
                f'Actualización terminada. Productos actualizados: {actualizados}.'
            )

            return redirect('inventario:inventario_producto_list')

        except Exception as error:
            messages.error(
                request,
                f'Error al actualizar productos: {error}'
            )

    return render(
        request,
        'inventario/actualizar_productos_excel.html'
    )