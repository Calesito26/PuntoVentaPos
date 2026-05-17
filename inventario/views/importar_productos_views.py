from decimal import Decimal

import openpyxl

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render

from inventario.models import Sede
from inventario.models import StockBodega
from productos.models import Categoria
from productos.models import Producto
from usuarios.decorators import administrador_required


def generar_codigo_producto():
    ultimo = Producto.objects.order_by('id').last()

    if not ultimo:
        numero = 1
    else:
        numero = ultimo.id + 1

    return f'PROD-{numero:06d}'


@administrador_required
def importar_productos_excel(request):

    if request.method == 'POST':

        archivo = request.FILES.get('archivo_excel')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('inventario:importar_productos_excel')

        try:
            workbook = openpyxl.load_workbook(archivo)
            hoja = workbook.active

            creados = 0
            actualizados = 0
            errores = []

            for fila in range(2, hoja.max_row + 1):

                nombre = hoja.cell(fila, 1).value
                categoria_id = hoja.cell(fila, 2).value
                precio_compra = hoja.cell(fila, 3).value or 0
                precio_venta = hoja.cell(fila, 4).value or 0
                imagen_url = hoja.cell(fila, 5).value
                bodega_id = hoja.cell(fila, 6).value
                stock = hoja.cell(fila, 7).value or 0
                stock_minimo = hoja.cell(fila, 8).value or 5
                activo = hoja.cell(fila, 9).value
                es_insumo_excel = str(hoja.cell(fila, 10).value or '').strip().upper()
                es_insumo = es_insumo_excel == 'SI'
                receta_servicio_combo = str(hoja.cell(fila, 11).value or 'NINGUNO').strip().upper()
                tipos_validos = ['NINGUNO','RECETA','SERVICIO','COMBO']

                if receta_servicio_combo not in tipos_validos:
                    receta_servicio_combo = 'NINGUNO'


                if not nombre:
                    continue

                try:
                    categoria = Categoria.objects.get(id=categoria_id)
                except Categoria.DoesNotExist:
                    errores.append(
                        f'Fila {fila}: categoría no existe.'
                    )
                    continue

                try:
                    sede = Sede.objects.get(id=bodega_id)
                except Sede.DoesNotExist:
                    errores.append(
                        f'Fila {fila}: bodega no existe.'
                    )
                    continue

                if str(activo).upper() in ['NO', '0', 'FALSE', 'INACTIVO']:
                    activo_producto = False
                else:
                    activo_producto = True

                producto = Producto.objects.filter(
                    nombre__iexact=nombre
                ).first()

                if producto:
                    producto.categoria = categoria
                    producto.precio_compra = Decimal(str(precio_compra))
                    producto.precio_venta = Decimal(str(precio_venta))
                    producto.imagen_url = imagen_url
                    producto.es_insumo = es_insumo
                    producto.receta_servicio_combo = receta_servicio_combo
                    producto.activo = activo_producto
                    producto.save()
                    actualizados += 1
                else:
                    producto = Producto.objects.create(
                        codigo=generar_codigo_producto(),
                        nombre=nombre,
                        categoria=categoria,
                        precio_compra=Decimal(str(precio_compra)),
                        precio_venta=Decimal(str(precio_venta)),
                        imagen_url=imagen_url,
                        es_insumo=es_insumo,
                        receta_servicio_combo=receta_servicio_combo,
                        activo=activo_producto
                    )
                    creados += 1

                StockBodega.objects.update_or_create(
                    sede=sede,
                    producto=producto,
                    defaults={
                        'stock': Decimal(str(stock)),
                        'stock_minimo': Decimal(str(stock_minimo)),
                        'activo': activo_producto
                    }
                )

            if errores:
                for error in errores:
                    messages.warning(request, error)

            messages.success(
                request,
                f'Importación terminada. Creados: {creados}. Actualizados: {actualizados}.'
            )

            return redirect('inventario:inventario_producto_list')

        except Exception as error:
            messages.error(
                request,
                f'Error al importar archivo: {error}'
            )

    return render(
        request,
        'inventario/importar_productos_excel.html'
    )


@administrador_required
def descargar_formato_productos(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = 'Productos'

    columnas = [
        'nombre',
        'categoria_id',
        'precio_compra',
        'precio_venta',
        'imagen_url',
        'bodega_id',
        'stock',
        'stock_minimo',
        'activo',
        'es_insumo',
        'receta_servicio_combo'
    ]

    hoja.append(columnas)

    hoja.append([
        'LECHE GLORIA EN TARRO',
        1,
        3.50,
        4.20,
        'https://plazavea.vteximg.com.br/arquivos/ids/261327/leche-gloria.jpg',
        1,
        20,
        5,
        'SI',
        'NO',
        'NINGUNO'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=formato_importar_productos.xlsx'

    workbook.save(response)

    return response