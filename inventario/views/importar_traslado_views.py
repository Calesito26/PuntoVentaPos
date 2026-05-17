from decimal import Decimal

import openpyxl

from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render

from inventario.models import Sede
from inventario.models import StockBodega
from inventario.models import TrasladoBodega
from productos.models import Producto
from usuarios.decorators import administrador_required


@administrador_required
def descargar_formato_traslados(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = 'Traslados'

    hoja.append([
        'bodega_origen_id',
        'bodega_destino_id',
        'codigo_producto',
        'cantidad',
        'observacion',
    ])

    hoja.append([
        1,
        2,
        'PROD-000001',
        5,
        'Traslado de prueba',
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=formato_importar_traslados.xlsx'

    workbook.save(response)

    return response


@administrador_required
@transaction.atomic
def importar_traslados_excel(request):

    if request.method == 'POST':

        archivo = request.FILES.get('archivo_excel')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('inventario:importar_traslados_excel')

        workbook = openpyxl.load_workbook(archivo)
        hoja = workbook.active

        importados = 0
        errores = []

        for fila in range(2, hoja.max_row + 1):

            bodega_origen_id = hoja.cell(fila, 1).value
            bodega_destino_id = hoja.cell(fila, 2).value
            codigo_producto = hoja.cell(fila, 3).value
            cantidad = hoja.cell(fila, 4).value
            observacion = hoja.cell(fila, 5).value or ''

            if not bodega_origen_id and not bodega_destino_id and not codigo_producto:
                continue

            sede_origen = Sede.objects.filter(id=bodega_origen_id).first()
            sede_destino = Sede.objects.filter(id=bodega_destino_id).first()
            producto = Producto.objects.filter(codigo=str(codigo_producto).strip()).first()

            if not sede_origen:
                errores.append(f'Fila {fila}: bodega origen no existe.')
                continue

            if not sede_destino:
                errores.append(f'Fila {fila}: bodega destino no existe.')
                continue

            if sede_origen == sede_destino:
                errores.append(f'Fila {fila}: origen y destino no pueden ser iguales.')
                continue

            if not producto:
                errores.append(f'Fila {fila}: producto no existe.')
                continue

            cantidad = Decimal(str(cantidad or 0))

            if cantidad <= 0:
                errores.append(f'Fila {fila}: cantidad inválida.')
                continue

            stock_origen = StockBodega.objects.filter(
                sede=sede_origen,
                producto=producto,
                activo=True
            ).first()

            if not stock_origen:
                errores.append(f'Fila {fila}: producto sin stock en bodega origen.')
                continue

            if stock_origen.stock < cantidad:
                errores.append(f'Fila {fila}: stock insuficiente para {producto.nombre}.')
                continue

            stock_destino, creado = StockBodega.objects.get_or_create(
                sede=sede_destino,
                producto=producto,
                defaults={
                    'stock': Decimal('0.00'),
                    'stock_minimo': stock_origen.stock_minimo,
                    'activo': True,
                }
            )

            cantidad_actual = stock_origen.stock

            stock_origen.stock -= cantidad
            stock_origen.save()

            stock_destino.stock += cantidad
            stock_destino.save()

            TrasladoBodega.objects.create(
                sede_origen=sede_origen,
                sede_destino=sede_destino,
                producto=producto,
                cantidad_actual=cantidad_actual,
                cantidad_traslado=cantidad,
                cantidad_final=stock_origen.stock,
                valor_traslado=cantidad * producto.precio_compra,
                responsable=request.user,
                observacion=observacion,
                estado='REALIZADO',
            )

            importados += 1

        for error in errores:
            messages.warning(request, error)

        messages.success(
            request,
            f'Importación terminada. Traslados importados: {importados}.'
        )

        return redirect('inventario:traslado_list')

    return render(
        request,
        'inventario/importar_traslados_excel.html'
    )