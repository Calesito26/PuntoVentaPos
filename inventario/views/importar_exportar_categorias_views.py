import openpyxl

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render

from productos.models import Categoria
from usuarios.decorators import administrador_required


@administrador_required
def importar_categorias_excel(request):

    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('inventario:importar_categorias_excel')

        workbook = openpyxl.load_workbook(archivo)
        hoja = workbook.active

        creadas = 0
        actualizadas = 0

        for fila in range(2, hoja.max_row + 1):
            nombre = hoja.cell(fila, 1).value
            activo = hoja.cell(fila, 2).value

            if not nombre:
                continue

            activo_categoria = str(activo or 'SI').strip().upper() not in [
                'NO',
                '0',
                'FALSE',
                'INACTIVO'
            ]

            categoria, creada = Categoria.objects.update_or_create(
                nombre=str(nombre).strip(),
                defaults={
                    'activo': activo_categoria
                }
            )

            if creada:
                creadas += 1
            else:
                actualizadas += 1

        messages.success(
            request,
            f'Importación terminada. Creadas: {creadas}. Actualizadas: {actualizadas}.'
        )

        return redirect('inventario:categoria_inventario_list')

    return render(
        request,
        'inventario/importar_categorias_excel.html'
    )


@administrador_required
def descargar_formato_categorias(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = 'Categorias'

    hoja.append([
        'nombre',
        'activo'
    ])

    hoja.append([
        'LACTEOS',
        'SI'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=formato_importar_categorias.xlsx'

    workbook.save(response)

    return response


@administrador_required
def exportar_categorias_excel(request):
    categorias = Categoria.objects.all().order_by('id')

    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = 'Categorias'

    hoja.append([
        'Item',
        'ID Categoría',
        'Nombre',
        'Estado'
    ])

    for index, categoria in enumerate(categorias, start=1):
        hoja.append([
            index,
            categoria.id,
            categoria.nombre,
            'Activo' if categoria.activo else 'Inactivo'
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'

    workbook.save(response)

    return response


@administrador_required
def exportar_categorias_pdf(request):
    categorias = Categoria.objects.all().order_by('id')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=categorias.pdf'

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    y = height - 40

    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(40, y, 'Lista de categorías')
    y -= 30

    pdf.setFont('Helvetica-Bold', 9)
    pdf.drawString(40, y, 'Item')
    pdf.drawString(90, y, 'ID')
    pdf.drawString(150, y, 'Nombre')
    pdf.drawString(400, y, 'Estado')
    y -= 15

    pdf.setFont('Helvetica', 9)

    for index, categoria in enumerate(categorias, start=1):
        if y < 40:
            pdf.showPage()
            y = height - 40
            pdf.setFont('Helvetica', 9)

        pdf.drawString(40, y, str(index))
        pdf.drawString(90, y, str(categoria.id))
        pdf.drawString(150, y, categoria.nombre[:35])
        pdf.drawString(400, y, 'Activo' if categoria.activo else 'Inactivo')

        y -= 15

    pdf.save()

    return response